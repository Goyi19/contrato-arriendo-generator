"""
Script de Actualización Automática de Estados - Carmen 339
===========================================================

Este script actualiza automáticamente los estados de los departamentos
desde el archivo del CRM hacia el archivo de estatus visual.

Flujo:
1. Lee el archivo del CRM (Productos Camen 339.xlsx)
2. Extrae número de departamento y estado
3. Calcula la posición en la matriz del edificio
4. Actualiza el archivo de estatus visual
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime

# ============================================
# CONFIGURACIÓN
# ============================================

# Detectar la carpeta donde está el script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Archivos de entrada y salida
ARCHIVO_CRM = os.path.join(SCRIPT_DIR, "Productos Camen 339.xls")
ARCHIVO_ESTATUS = os.path.join(SCRIPT_DIR, "Estatus Carmen 339 v4 - JF 141125 - copia.xlsx")
LOG_PATH = os.path.join(SCRIPT_DIR, "logs")

# Configuración de hojas
HOJA_CRM = "Productos"
HOJA_ESTATUS = "Carmen 339"

# Configuración de columnas en el CRM
FILA_INICIO_CRM = 8  # Primera fila de datos (C8, K8)
COL_NOMBRE_CRM = "C"  # Columna con número de departamento
COL_ESTADO_CRM = "K"  # Columna con estado

# Configuración de la matriz en el archivo de estatus
FILA_PISO_9 = 13  # Fila donde empieza el piso 9
FILA_PISO_1 = 21  # Fila donde está el piso 1
COL_COLUMNA_1 = "C"  # Columna donde empieza la columna 1 del edificio

# ============================================
# FUNCIONES AUXILIARES
# ============================================

def crear_directorio_logs():
    """Crea el directorio de logs si no existe"""
    if not os.path.exists(LOG_PATH):
        os.makedirs(LOG_PATH)

def registrar_log(mensaje):
    """Registra mensajes en un archivo de log"""
    crear_directorio_logs()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_file = os.path.join(LOG_PATH, f"actualizacion_{datetime.now().strftime('%Y%m')}.txt")
    
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {mensaje}\n")
    
    print(f"[{timestamp}] {mensaje}")

def parsear_numero_departamento(numero):
    """
    Extrae piso y columna del número de departamento.
    Ejemplo: 107 → Piso 1, Columna 7
             926 → Piso 9, Columna 26
    """
    numero_str = str(int(numero))  # Convertir a string sin decimales
    
    if len(numero_str) == 3:
        # Formato: XYZ → Piso X, Columna YZ
        piso = int(numero_str[0])
        columna = int(numero_str[1:])
    elif len(numero_str) == 2:
        # Formato: XY → Piso X, Columna Y (raro pero posible)
        piso = int(numero_str[0])
        columna = int(numero_str[1])
    else:
        raise ValueError(f"Formato de departamento no reconocido: {numero}")
    
    return piso, columna

def calcular_celda_destino(piso, columna):
    """
    Calcula la celda de destino en la matriz del edificio.
    
    Matriz:
    - Filas: Piso 9 en fila 13, Piso 1 en fila 21
    - Columnas: Columna 1 en columna C (índice 3), hasta columna AB
    """
    # Calcular fila (piso 9 arriba, piso 1 abajo)
    fila = FILA_PISO_9 + (9 - piso)
    
    # Calcular columna (empieza en C que es la columna 3 en Excel)
    # Columna 1 del edificio = Columna C de Excel (índice 3)
    col_excel = 2 + columna  # B=2, C=3, entonces C=2+1
    
    return fila, col_excel

def columna_numero_a_letra(num):
    """Convierte número de columna (1-based) a letra de Excel"""
    resultado = ""
    while num > 0:
        num -= 1
        resultado = chr(65 + num % 26) + resultado
        num //= 26
    return resultado

# ============================================
# FUNCIÓN PRINCIPAL
# ============================================

def actualizar_estados():
    """Función principal que ejecuta la actualización"""
    
    print("\n" + "="*70)
    print(" "*15 + "ACTUALIZACIÓN DE ESTADOS - CARMEN 339")
    print("="*70 + "\n")
    
    try:
        # ============================================
        # PASO 1: Verificar que existan los archivos
        # ============================================
        
        registrar_log("🔍 Verificando archivos...")
        
        if not os.path.exists(ARCHIVO_CRM):
            registrar_log(f"❌ ERROR: No se encontró el archivo del CRM: {ARCHIVO_CRM}")
            return False
        
        if not os.path.exists(ARCHIVO_ESTATUS):
            registrar_log(f"❌ ERROR: No se encontró el archivo de estatus: {ARCHIVO_ESTATUS}")
            return False
        
        registrar_log("✅ Archivos encontrados")
        
        # ============================================
        # PASO 2: Leer datos del CRM
        # ============================================
        
        registrar_log("📂 Leyendo archivo del CRM...")
        
        # Leer Excel del CRM
        df_crm = pd.read_excel(ARCHIVO_CRM, sheet_name=HOJA_CRM, header=None)
        
        # Extraer columnas de interés (C y K, que son índices 2 y 10 en pandas)
        # Desde fila 8 en Excel = índice 7 en pandas
        col_nombres = df_crm.iloc[FILA_INICIO_CRM-1:, 2]  # Columna C
        col_estados = df_crm.iloc[FILA_INICIO_CRM-1:, 10]  # Columna K
        
        # Crear diccionario: {numero_depto: estado}
        estados_departamentos = {}
        filas_procesadas = 0
        
        for nombre, estado in zip(col_nombres, col_estados):
            # Saltar valores vacíos
            if pd.isna(nombre) or pd.isna(estado):
                continue
            
            # Saltar si no es un número
            try:
                numero_depto = int(float(nombre))
                estado_limpio = str(estado).strip()
                
                # Mapear "Entregado" a "Escriturado"
                if estado_limpio.upper() == "ENTREGADO":
                    estado_limpio = "Escriturado"
                
                estados_departamentos[numero_depto] = estado_limpio
                filas_procesadas += 1
            except (ValueError, TypeError):
                continue
        
        registrar_log(f"✅ Leídos {filas_procesadas} departamentos del CRM")
        
        if filas_procesadas == 0:
            registrar_log("⚠️ ADVERTENCIA: No se encontraron departamentos en el CRM")
            return False
        
        # ============================================
        # PASO 3: Abrir archivo de estatus con openpyxl
        # ============================================
        
        registrar_log("📝 Abriendo archivo de estatus...")
        
        wb = openpyxl.load_workbook(ARCHIVO_ESTATUS)
        ws = wb[HOJA_ESTATUS]
        
        # ============================================
        # PASO 4: Actualizar estados en la matriz
        # ============================================
        
        registrar_log("🔄 Actualizando estados en la matriz...")
        
        actualizados = 0
        errores = 0
        
        for numero_depto, estado in estados_departamentos.items():
            try:
                # Parsear número de departamento
                piso, columna = parsear_numero_departamento(numero_depto)
                
                # Validar rangos
                if piso < 1 or piso > 9:
                    registrar_log(f"⚠️ Piso inválido para depto {numero_depto}: piso {piso}")
                    errores += 1
                    continue
                
                if columna < 1 or columna > 26:
                    registrar_log(f"⚠️ Columna inválida para depto {numero_depto}: columna {columna}")
                    errores += 1
                    continue
                
                # Calcular celda destino
                fila, col_num = calcular_celda_destino(piso, columna)
                col_letra = columna_numero_a_letra(col_num)
                celda_ref = f"{col_letra}{fila}"
                
              # Obtener celda
                celda = ws[celda_ref]
                
                # Leer valor actual de la celda
                valor_actual = celda.value
                
                # Actualizar solo si NO es "Disponible" Y si el valor cambió
                if estado.strip().upper() != "DISPONIBLE":
                    if str(valor_actual).strip() != estado.strip():
                        celda.value = estado
                        celda.alignment = Alignment(horizontal='center', vertical='center')
                        actualizados += 1
                    else:
                        # Ya tiene el valor correcto, no hacer nada
                        pass
                else:
                    # No actualizar si está disponible (mantiene el valor UF)
                    registrar_log(f"⏭️ Depto {numero_depto} está Disponible - Manteniendo valor actual")
            except Exception as e:
                registrar_log(f"❌ Error procesando depto {numero_depto}: {str(e)}")
                errores += 1
        
        registrar_log(f"✅ Actualizados en matriz: {actualizados} departamentos")
        if errores > 0:
            registrar_log(f"⚠️ Errores en matriz: {errores} departamentos")
        
        # ============================================
        # PASO 4B: Actualizar tabla de comentarios
        # ============================================
        
        registrar_log("🔄 Actualizando tabla de comentarios...")
        
        FILA_INICIO_COMENTARIOS = 4  # Primera fila de datos (AE4, AG4)
        COL_NRO_UNIDAD = "AE"
        COL_ESTATUS_COMENTARIOS = "AG"
        
        actualizados_comentarios = 0
        errores_comentarios = 0
        
        fila_actual = FILA_INICIO_COMENTARIOS
        
        while True:
            # Leer número de unidad
            celda_unidad = ws[f"{COL_NRO_UNIDAD}{fila_actual}"]
            valor_unidad = celda_unidad.value
            
            # Si la celda está vacía o no es un número, salir del loop
            if valor_unidad is None or valor_unidad == "":
                break
            
            try:
                numero_depto = int(float(valor_unidad))
            except (ValueError, TypeError):
                # No es un número, probablemente llegamos al siguiente bloque
                break
            
            # Si este departamento está en nuestro diccionario, actualizar
            if numero_depto in estados_departamentos:
                estado_nuevo = estados_departamentos[numero_depto]
                
                # Obtener celda de estado
                celda_estado = ws[f"{COL_ESTATUS_COMENTARIOS}{fila_actual}"]
                valor_actual = celda_estado.value
                
                # Actualizar solo si NO es "Disponible" Y si el valor cambió
                if estado_nuevo.strip().upper() != "DISPONIBLE":
                    if str(valor_actual).strip() != estado_nuevo.strip():
                        celda_estado.value = estado_nuevo
                        celda_estado.alignment = Alignment(horizontal='center', vertical='center')
                        actualizados_comentarios += 1
                        registrar_log(f"   ✅ Comentarios: Depto {numero_depto} → {estado_nuevo}")
            
            fila_actual += 1
            
            # Límite de seguridad para evitar loop infinito
            if fila_actual > 1000:
                registrar_log("⚠️ Límite de filas alcanzado en tabla de comentarios")
                break
        
        registrar_log(f"✅ Actualizados en comentarios: {actualizados_comentarios} departamentos")
        
        # ============================================
        # PASO 5: Guardar archivo
        # ============================================
        
        registrar_log("💾 Guardando archivo...")
        
        # Crear backup del archivo original (opcional)
        # timestamp_backup = datetime.now().strftime("%Y%m%d_%H%M%S")
        # archivo_backup = ARCHIVO_ESTATUS.replace(".xlsx", f"_backup_{timestamp_backup}.xlsx")
        # wb.save(archivo_backup)
        # registrar_log(f"📦 Backup creado: {archivo_backup}")
        
        # Guardar archivo actualizado
        wb.save(ARCHIVO_ESTATUS)
        wb.close()
        
        registrar_log("✅ Archivo guardado exitosamente")
        
        # ============================================
        # RESUMEN FINAL
        # ============================================
        
        print("\n" + "="*70)
        print(" "*25 + "RESUMEN DE ACTUALIZACIÓN")
        print("="*70)
        print(f"\n📊 Departamentos procesados del CRM: {filas_procesadas}")
        print(f"✅ Actualizados en matriz: {actualizados}")
        print(f"✅ Actualizados en comentarios: {actualizados_comentarios}")
        if errores > 0:
            print(f"⚠️ Con errores: {errores}")
        print(f"\n📁 Archivo actualizado: {ARCHIVO_ESTATUS}")
        print("="*70 + "\n")
        
        return True
        
    except FileNotFoundError as e:
        registrar_log(f"❌ ERROR: Archivo no encontrado - {str(e)}")
        return False
    
    except Exception as e:
        registrar_log(f"❌ ERROR INESPERADO: {str(e)}")
        import traceback
        registrar_log(traceback.format_exc())
        return False

# ============================================
# EJECUCIÓN PRINCIPAL
# ============================================

if __name__ == "__main__":
    try:
        resultado = actualizar_estados()
        
        if resultado:
            print("┌" + "─"*68 + "┐")
            print("│" + " "*20 + "✅ ACTUALIZACIÓN EXITOSA" + " "*24 + "│")
            print("├" + "─"*68 + "┤")
            print("│  El archivo de estatus ha sido actualizado correctamente.       │")
            print("│  Puedes abrirlo y verificar los cambios.                        │")
            print("└" + "─"*68 + "┘")
        else:
            print("┌" + "─"*68 + "┐")
            print("│" + " "*17 + "❌ ACTUALIZACIÓN CON ERRORES" + " "*23 + "│")
            print("├" + "─"*68 + "┤")
            print("│  Revisa los errores mostrados arriba.                           │")
            print("│  Verifica que los archivos existan y tengan el formato correcto.│")
            print("└" + "─"*68 + "┘")
        
        # Pausa para que se pueda leer el resultado
        print("\n")
        input("Presiona ENTER para cerrar...")
        
    except Exception as e:
        print("\n" + "="*70)
        print("❌ ERROR CRÍTICO AL EJECUTAR EL SCRIPT:")
        print(f"   {str(e)}")
        print("="*70)
        print("\n⚠️ Contacta al equipo de soporte técnico con este mensaje de error\n")
        input("Presiona ENTER para cerrar...")
