# -*- coding: utf-8 -*-
"""
Created on Wed Jan 21 17:40:54 2026

@author: Sofi
"""


"""
Script de Actualización Automática de Estados - Tarapacá 1140
==============================================================

Este script actualiza automáticamente los estados de los departamentos
desde el archivo del CRM hacia el archivo de estatus visual.

Flujo:
1. Lee el archivo del CRM (Productos_TARAPACA_1140.xlsx)
2. Extrae número de departamento y estado
3. Calcula la posición en la matriz del edificio
4. Actualiza el archivo de estatus visual
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime

# ============================================
# CONFIGURACIÓN
# ============================================

# Detectar la carpeta donde está el script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Archivos de entrada y salida
ARCHIVO_CRM = os.path.join(SCRIPT_DIR, "Productos_TARAPACA_1140.xls")
ARCHIVO_ESTATUS = os.path.join(SCRIPT_DIR, "20251029 Estatus Tarapacá JF.xlsx")
LOG_PATH = os.path.join(SCRIPT_DIR, "logs")

# Configuración de hojas
HOJA_CRM = "Productos"
HOJA_ESTATUS = "Tarapacá 1140"

# Configuración de columnas en el CRM
FILA_INICIO_CRM = 8  # Primera fila de datos (C8, K8)
COL_NOMBRE_CRM = "C"  # Columna con número de departamento
COL_ESTADO_CRM = "K"  # Columna con estado

# Configuración de la matriz en el archivo de estatus
FILA_PISO_9 = 10  # Fila donde empieza el piso 9 (B10)
FILA_PISO_1 = 18  # Fila donde está el piso 1 (B18)
COL_COLUMNA_1 = "C"  # Columna donde empieza la columna 1 del edificio (C4)

# Configuración del bloque de comentarios
FILA_INICIO_COMENTARIOS = 4  # Primera fila de datos (V4, X4)
COL_NRO_UNIDAD = "V"
COL_ESTATUS_COMENTARIOS = "X"

# ============================================
# FUNCIONES AUXILIARES
# ============================================

def crear_directorio_logs():
    """Crea el directorio de logs si no existe"""
    if not os.path.exists(LOG_PATH):
        os.makedirs(LOG_PATH)

def registrar_log(mensaje):
    """Registra mensajes en un archivo de log"""
    crear_directorio_logs()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_file = os.path.join(LOG_PATH, f"actualizacion_{datetime.now().strftime('%Y%m')}.txt")
    
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {mensaje}\n")
    
    print(f"[{timestamp}] {mensaje}")

def parsear_numero_departamento(numero):
    """
    Extrae piso y columna del número de departamento.
    Ejemplo: 107 → Piso 1, Columna 7
             917 → Piso 9, Columna 17
    """
    numero_str = str(int(numero))  # Convertir a string sin decimales
    
    if len(numero_str) == 3:
        # Formato: XYZ → Piso X, Columna YZ
        piso = int(numero_str[0])
        columna = int(numero_str[1:])
    elif len(numero_str) == 2:
        # Formato: XY → Piso X, Columna Y
        piso = int(numero_str[0])
        columna = int(numero_str[1])
    else:
        raise ValueError(f"Formato de departamento no reconocido: {numero}")
    
    return piso, columna

def calcular_celda_destino(piso, columna):
    """
    Calcula la celda de destino en la matriz del edificio.
    
    Matriz:
    - Filas: Piso 9 en fila 10, Piso 1 en fila 18
    - Columnas: Columna 1 en columna C (índice 3), hasta columna S (17 columnas)
    """
    # Calcular fila (piso 9 arriba, piso 1 abajo)
    fila = FILA_PISO_9 + (9 - piso)
    
    # Calcular columna (empieza en C que es la columna 3 en Excel)
    # Columna 1 del edificio = Columna C de Excel (índice 3)
    col_excel = 2 + columna  # C=3, entonces C=2+1
    
    return fila, col_excel

def columna_numero_a_letra(num):
    """Convierte número de columna (1-based) a letra de Excel"""
    resultado = ""
    while num > 0:
        num -= 1
        resultado = chr(65 + num % 26) + resultado
        num //= 26
    return resultado

# ============================================
# FUNCIÓN PRINCIPAL
# ============================================

def actualizar_estados():
    """Función principal que ejecuta la actualización"""
    
    print("\n" + "="*70)
    print(" "*15 + "ACTUALIZACIÓN DE ESTADOS - TARAPACÁ 1140")
    print("="*70 + "\n")
    
    try:
        # ============================================
        # PASO 1: Verificar que existan los archivos
        # ============================================
        
        registrar_log("🔍 Verificando archivos...")
        
        if not os.path.exists(ARCHIVO_CRM):
            registrar_log(f"❌ ERROR: No se encontró el archivo del CRM: {ARCHIVO_CRM}")
            return False
        
        if not os.path.exists(ARCHIVO_ESTATUS):
            registrar_log(f"❌ ERROR: No se encontró el archivo de estatus: {ARCHIVO_ESTATUS}")
            return False
        
        registrar_log("✅ Archivos encontrados")
        
        # ============================================
        # PASO 2: Leer datos del CRM
        # ============================================
        
        registrar_log("📂 Leyendo archivo del CRM...")
        
        # Leer Excel del CRM
        df_crm = pd.read_excel(ARCHIVO_CRM, sheet_name=HOJA_CRM, header=None)
        
        # Extraer columnas de interés (C y K, que son índices 2 y 10 en pandas)
        # Desde fila 8 en Excel = índice 7 en pandas
        col_nombres = df_crm.iloc[FILA_INICIO_CRM-1:, 2]  # Columna C
        col_estados = df_crm.iloc[FILA_INICIO_CRM-1:, 10]  # Columna K
        
        # Crear diccionario: {numero_depto: estado}
        estados_departamentos = {}
        filas_procesadas = 0
        
        for nombre, estado in zip(col_nombres, col_estados):
            # Saltar valores vacíos
            if pd.isna(nombre) or pd.isna(estado):
                continue
            
            # Extraer número del departamento
            try:
                nombre_str = str(nombre).strip()
                
                # Si tiene formato "Departamento XXX", extraer solo el número
                if "departamento" in nombre_str.lower():
                    # Extraer todos los dígitos
                    import re
                    numeros = re.findall(r'\d+', nombre_str)
                    if numeros:
                        numero_depto = int(numeros[0])
                    else:
                        continue
                else:
                    # Si es solo un número
                    numero_depto = int(float(nombre_str))
                
                estado_limpio = str(estado).strip()
                
                # Mapear "Entregado" a "Escriturado"
                if estado_limpio.upper() == "ENTREGADO":
                    estado_limpio = "Escriturado"
                
                estados_departamentos[numero_depto] = estado_limpio
                filas_procesadas += 1
            except (ValueError, TypeError):
                continue
        
        registrar_log(f"✅ Leídos {filas_procesadas} departamentos del CRM")
        
        if filas_procesadas == 0:
            registrar_log("⚠️ ADVERTENCIA: No se encontraron departamentos en el CRM")
            return False
        
        # ============================================
        # PASO 3: Abrir archivo de estatus con openpyxl
        # ============================================
        
        registrar_log("📖 Abriendo archivo de estatus...")
        
        wb = openpyxl.load_workbook(ARCHIVO_ESTATUS)
        ws = wb[HOJA_ESTATUS]
        
        # ============================================
        # PASO 4: Actualizar estados en la matriz
        # ============================================
        
        registrar_log("🔄 Actualizando estados en la matriz...")
        
        actualizados = 0
        errores = 0
        
        # Diccionario para guardar valores anteriores (ANTES de actualizar)
        valores_anteriores_matriz = {}
        
        # Lista para guardar los cambios realizados en la matriz
        cambios_matriz = []
        
        for numero_depto, estado in estados_departamentos.items():
            try:
                # Parsear número de departamento
                piso, columna = parsear_numero_departamento(numero_depto)
                
                # Validar rangos
                if piso < 1 or piso > 9:
                    registrar_log(f"⚠️ Piso inválido para depto {numero_depto}: piso {piso}")
                    errores += 1
                    continue
                
                if columna < 1 or columna > 17:
                    registrar_log(f"⚠️ Columna inválida para depto {numero_depto}: columna {columna}")
                    errores += 1
                    continue
                
                # Calcular celda destino
                fila, col_num = calcular_celda_destino(piso, columna)
                col_letra = columna_numero_a_letra(col_num)
                celda_ref = f"{col_letra}{fila}"
                
                # Obtener celda
                celda = ws[celda_ref]
                
                # Leer valor actual de la celda
                valor_actual = celda.value
                
                # GUARDAR valor anterior para usarlo después
                valores_anteriores_matriz[numero_depto] = valor_actual
                
                # NO actualizar si es SDV (Sala de Ventas) o Piloto - mantener tal cual
                if valor_actual and str(valor_actual).strip().upper() in ["SDV", "PILOTO"]:
                    registrar_log(f"   ⭐ Depto {numero_depto} es '{valor_actual}' - Manteniendo sin cambios")
                    continue
                
                # Actualizar solo si NO es "Disponible" Y si el valor cambió
                if estado.strip().upper() != "DISPONIBLE":
                    if str(valor_actual).strip() != estado.strip():
                        celda.value = estado
                        celda.alignment = Alignment(horizontal='center', vertical='center')
                        actualizados += 1
                        # Guardar el cambio para el resumen final
                        cambios_matriz.append((numero_depto, valor_actual, estado))
                        # Mostrar cambio detallado
                        registrar_log(f"   🔄 Depto {numero_depto}: '{valor_actual}' → '{estado}'")
                    else:
                        # Ya tiene el valor correcto, no hacer nada
                        pass
                else:
                    # No actualizar si está disponible (mantiene el valor UF)
                    registrar_log(f"   ⭐ Depto {numero_depto} está Disponible - Manteniendo '{valor_actual}'")

            except Exception as e:
                registrar_log(f"❌ Error procesando depto {numero_depto}: {str(e)}")
                errores += 1
        
        registrar_log(f"✅ Actualizados en matriz: {actualizados} departamentos")
        if errores > 0:
            registrar_log(f"⚠️ Errores en matriz: {errores} departamentos")
        
        # ============================================
        # PASO 4B: Actualizar tabla de comentarios
        # ============================================
        
        registrar_log("🔄 Actualizando tabla de comentarios...")
        
        actualizados_comentarios = 0
        errores_comentarios = 0
        
        fila_actual = FILA_INICIO_COMENTARIOS
        ultima_fila_con_datos = FILA_INICIO_COMENTARIOS - 1
        
        # ==========================================
        # PARTE 1: Actualizar departamentos existentes
        # ==========================================
        
        while True:
            # Leer número de unidad
            celda_unidad = ws[f"{COL_NRO_UNIDAD}{fila_actual}"]
            valor_unidad = celda_unidad.value
            
            # Si la celda está vacía o no es un número, salir del loop
            if valor_unidad is None or valor_unidad == "":
                ultima_fila_con_datos = fila_actual - 1  # Guardar última fila con datos
                break
            
            try:
                numero_depto = int(float(valor_unidad))
            except (ValueError, TypeError):
                # No es un número, probablemente llegamos al siguiente bloque
                ultima_fila_con_datos = fila_actual - 1
                break
            
            # Si este departamento está en nuestro diccionario, actualizar
            if numero_depto in estados_departamentos:
                estado_nuevo = estados_departamentos[numero_depto]
                
                # Obtener celda de estado
                celda_estado = ws[f"{COL_ESTATUS_COMENTARIOS}{fila_actual}"]
                valor_actual = celda_estado.value
                
                # Actualizar solo si NO es "Disponible" Y si el valor cambió
                if estado_nuevo.strip().upper() != "DISPONIBLE":
                    if str(valor_actual).strip() != estado_nuevo.strip():
                        celda_estado.value = estado_nuevo
                        celda_estado.alignment = Alignment(horizontal='center', vertical='center')
                        actualizados_comentarios += 1
                        registrar_log(f"   ✅ Comentarios: Depto {numero_depto} → {estado_nuevo}")
            
            fila_actual += 1
            
            # Límite de seguridad para evitar loop infinito
            if fila_actual > 1000:
                registrar_log("⚠️ Límite de filas alcanzado en tabla de comentarios")
                ultima_fila_con_datos = fila_actual - 1
                break
        
        registrar_log(f"✅ Actualizados en comentarios: {actualizados_comentarios} departamentos")
        
        # ==========================================
        # PARTE 2: Agregar nuevos departamentos que cambiaron de "Disponible"
        # ==========================================
        
        # registrar_log("🔍 Buscando departamentos que cambiaron de Disponible...")
        
        # # Obtener lista de departamentos que ya están en la tabla de comentarios
        # deptos_ya_en_tabla = set()
        # for fila in range(FILA_INICIO_COMENTARIOS, ultima_fila_con_datos + 1):
        #     celda = ws[f"{COL_NRO_UNIDAD}{fila}"]
        #     if celda.value:
        #         try:
        #             deptos_ya_en_tabla.add(int(float(celda.value)))
        #         except:
        #             pass
        
        # # Buscar departamentos que pasaron de Disponible a otro estado
        # nuevos_a_agregar = []
        
        # for numero_depto, estado_nuevo in estados_departamentos.items():
        #     # Solo agregar si:
        #     # 1. NO está "Disponible"
        #     # 2. NO está "Escriturado" (que era "Entregado")
        #     # 3. NO está "Bloqueado"
        #     # 4. NO está ya en la tabla de comentarios
            
        #     estado_upper = estado_nuevo.strip().upper()
            
        #     if (estado_upper != "DISPONIBLE" and 
        #         estado_upper != "ESCRITURADO" and
        #         estado_upper != "BLOQUEADO" and
        #         numero_depto not in deptos_ya_en_tabla):
                
        #         # Obtener el valor anterior que guardamos ANTES de actualizar
        #         try:
        #             # Usar el valor que guardamos antes de actualizar
        #             if numero_depto in valores_anteriores_matriz:
        #                 valor_anterior = valores_anteriores_matriz[numero_depto]
        #             else:
        #                 # Si no está en el diccionario, buscarlo en la matriz
        #                 piso, columna = parsear_numero_departamento(numero_depto)
        #                 fila_matriz, col_num_matriz = calcular_celda_destino(piso, columna)
        #                 col_letra_matriz = columna_numero_a_letra(col_num_matriz)
        #                 celda_matriz_ref = f"{col_letra_matriz}{fila_matriz}"
        #                 celda_matriz = ws[celda_matriz_ref]
        #                 valor_anterior = celda_matriz.value
                    
        #             # Convertir valor_anterior a string para análisis
        #             valor_ant_str = str(valor_anterior) if valor_anterior else ""
                    
        #             # Detectar si era un valor UF (número)
        #             es_valor_uf = False
                    
        #             if valor_anterior is None or valor_anterior == "":
        #                 es_valor_uf = True
        #             elif isinstance(valor_anterior, (int, float)):
        #                 es_valor_uf = True
        #             elif "$" in valor_ant_str or "UF" in valor_ant_str.upper():
        #                 es_valor_uf = True
        #             else:
        #                 # Intentar convertir a número
        #                 try:
        #                     float(valor_ant_str.replace(",", "").replace(".", "").strip())
        #                     es_valor_uf = True
        #                 except:
        #                     es_valor_uf = False
                    
        #             if es_valor_uf:
        #                 # Estaba disponible (con valor UF), ahora cambió → agregar
        #                 nuevos_a_agregar.append((numero_depto, estado_nuevo))
                    
        #         except:
        #             # Si hay error parseando, mejor no agregarlo
        #             pass
        
        # # Agregar los nuevos departamentos
        # if nuevos_a_agregar:
        #     registrar_log(f"📝 Agregando {len(nuevos_a_agregar)} departamentos nuevos a la tabla de comentarios...")
            
        #     # Insertar filas necesarias ANTES de la primera celda vacía
        #     fila_insercion = ultima_fila_con_datos + 1
            
        #     for i, (numero_depto, estado_nuevo) in enumerate(nuevos_a_agregar):
        #         # Insertar una fila nueva
        #         ws.insert_rows(fila_insercion + i)
                
        #         # Escribir número de departamento
        #         celda_depto = ws[f"{COL_NRO_UNIDAD}{fila_insercion + i}"]
        #         celda_depto.value = numero_depto
        #         celda_depto.alignment = Alignment(horizontal='center', vertical='center')
                
        #         # Escribir estado
        #         celda_estado = ws[f"{COL_ESTATUS_COMENTARIOS}{fila_insercion + i}"]
        #         celda_estado.value = estado_nuevo
        #         celda_estado.alignment = Alignment(horizontal='center', vertical='center')
                
        #         registrar_log(f"   ➕ Agregado: Depto {numero_depto} → {estado_nuevo}")
            
        #     registrar_log(f"✅ Agregados {len(nuevos_a_agregar)} departamentos a comentarios")
        # else:
        #     registrar_log("ℹ️ No hay nuevos departamentos para agregar a comentarios")

        
        # ============================================
        # PASO 5: Guardar archivo
        # ============================================
        
        registrar_log("💾 Guardando archivo...")
        
        # Guardar archivo actualizado
        wb.save(ARCHIVO_ESTATUS)
        wb.close()
        
        registrar_log("✅ Archivo guardado exitosamente")
        
        # ============================================
        # RESUMEN FINAL
        # ============================================
        
        print("\n" + "="*70)
        print(" "*25 + "RESUMEN DE ACTUALIZACIÓN")
        print("="*70)
        print(f"\n📊 Departamentos procesados del CRM: {filas_procesadas}")
        print(f"✅ Actualizados en matriz: {actualizados}")
        print(f"✅ Actualizados en comentarios: {actualizados_comentarios}")
        if errores > 0:
            print(f"⚠️ Con errores: {errores}")
        print(f"\n📁 Archivo actualizado: {ARCHIVO_ESTATUS}")
        
        # Mostrar detalle de cambios en la matriz
        if cambios_matriz:
            print("\n" + "="*70)
            print(" "*20 + "DETALLE DE CAMBIOS EN MATRIZ")
            print("="*70)
            for depto, valor_ant, valor_nuevo in cambios_matriz:
                print(f"  🔄 Departamento {depto}: '{valor_ant}' → '{valor_nuevo}'")
        
        print("="*70 + "\n")
        
        return True
        
    except FileNotFoundError as e:
        registrar_log(f"❌ ERROR: Archivo no encontrado - {str(e)}")
        return False
    
    except Exception as e:
        registrar_log(f"❌ ERROR INESPERADO: {str(e)}")
        import traceback
        registrar_log(traceback.format_exc())
        return False

# ============================================
# EJECUCIÓN PRINCIPAL
# ============================================

if __name__ == "__main__":
    try:
        resultado = actualizar_estados()
        
        if resultado:
            print("┌" + "─"*68 + "┐")
            print("│" + " "*20 + "✅ ACTUALIZACIÓN EXITOSA" + " "*24 + "│")
            print("├" + "─"*68 + "┤")
            print("│  El archivo de estatus ha sido actualizado correctamente.       │")
            print("│  Puedes abrirlo y verificar los cambios.                        │")
            print("└" + "─"*68 + "┘")
        else:
            print("┌" + "─"*68 + "┐")
            print("│" + " "*17 + "❌ ACTUALIZACIÓN CON ERRORES" + " "*23 + "│")
            print("├" + "─"*68 + "┤")
            print("│  Revisa los errores mostrados arriba.                           │")
            print("│  Verifica que los archivos existan y tengan el formato correcto.│")
            print("└" + "─"*68 + "┘")
        
        # Pausa para que se pueda leer el resultado
        print("\n")
        input("Presiona ENTER para cerrar...")
        
    except Exception as e:
        print("\n" + "="*70)
        print("❌ ERROR CRÍTICO AL EJECUTAR EL SCRIPT:")
        print(f"   {str(e)}")
        print("="*70)
        print("\n⚠️ Contacta al equipo de soporte técnico con este mensaje de error\n")
        input("Presiona ENTER para cerrar...")