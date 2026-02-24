from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import io
import re
import traceback

app = Flask(__name__)
CORS(app)

# Configuración adaptada de los scripts originales
CONFIGS = {
    "Carmen 339": {
        "FILA_PISO_9": 13,
        "FILA_PISO_1": 21,
        "COL_COLUMNA_1": 3, # C
        "COL_NRO_UNIDAD_COM": "AE",
        "COL_ESTATUS_COM": "AG",
        "MAX_COL": 26,
        "MIN_PISO": 1,
        "MAX_PISO": 9
    },
    "Tarapacá 1140": {
        "FILA_PISO_9": 10,
        "FILA_PISO_1": 18,
        "COL_COLUMNA_1": 3, # C
        "COL_NRO_UNIDAD_COM": "V",
        "COL_ESTATUS_COM": "X",
        "MAX_COL": 17,
        "MIN_PISO": 1,
        "MAX_PISO": 9
    }
}

def parsear_numero_departamento(numero):
    """Extrae piso y columna del número de departamento."""
    try:
        numero_str = str(int(float(numero)))
        if len(numero_str) == 3:
            piso = int(numero_str[0])
            columna = int(numero_str[1:])
        elif len(numero_str) == 2:
            piso = int(numero_str[0])
            columna = int(numero_str[1])
        else:
            return None, None
        return piso, columna
    except:
        return None, None

@app.route('/api/update_status', methods=['POST'])
def update_status():
    try:
        if 'file_crm' not in request.files or 'file_estatus' not in request.files:
            return jsonify({"error": "Faltan archivos file_crm o file_estatus"}), 400

        file_crm = request.files['file_crm']
        file_estatus = request.files['file_estatus']

        # 1. Leer datos del CRM
        # Nota: xlrd es necesario para archivos .xls antiguos
        # Intentar leer con pandas
        try:
            # Forzar motor si es necesario, pero pandas suele detectarlo
            df_crm = pd.read_excel(file_crm, sheet_name="Productos", header=None)
        except Exception as e:
            return jsonify({"error": f"Error al leer el archivo CRM: {str(e)}"}), 400

        # Extraer nombres (col C=2) y estados (col K=10) desde fila 8 (index 7)
        col_nombres = df_crm.iloc[7:, 2]
        col_estados = df_crm.iloc[7:, 10]

        estados_departamentos = {}
        for nombre, estado in zip(col_nombres, col_estados):
            if pd.isna(nombre) or pd.isna(estado):
                continue
            
            try:
                nombre_str = str(nombre).strip()
                # Limpiar nombre de departamento si incluye texto
                if "departamento" in nombre_str.lower():
                    numeros = re.findall(r'\d+', nombre_str)
                    if numeros:
                        numero_depto = int(numeros[0])
                    else:
                        continue
                else:
                    numero_depto = int(float(nombre_str))
                
                estado_limpio = str(estado).strip()
                # Mapeo especial
                if estado_limpio.upper() == "ENTREGADO":
                    estado_limpio = "Escriturado"
                
                estados_departamentos[numero_depto] = estado_limpio
            except:
                continue

        if not estados_departamentos:
            return jsonify({"error": "No se encontraron departamentos válidos en el CRM"}), 400

        # 2. Abrir Plantilla de Estatus
        try:
            # Usar BytesIO para no guardar en disco
            file_estatus_bytes = io.BytesIO(file_estatus.read())
            wb = openpyxl.load_workbook(file_estatus_bytes)
        except Exception as e:
            return jsonify({"error": f"Error al abrir la plantilla de estatus: {str(e)}"}), 400

        # 3. Detectar Edificio/Configuración
        config = None
        ws = None
        found_sheet = ""
        
        for sheet_name in wb.sheetnames:
            if sheet_name in CONFIGS:
                config = CONFIGS[sheet_name]
                ws = wb[sheet_name]
                found_sheet = sheet_name
                break
        
        if not config:
            # Intentar búsqueda parcial si no hay match exacto
            for sheet_name in wb.sheetnames:
                for conf_name in CONFIGS:
                    if conf_name.lower() in sheet_name.lower():
                        config = CONFIGS[conf_name]
                        ws = wb[sheet_name]
                        found_sheet = sheet_name
                        break
                if config: break

        if not config:
            return jsonify({
                "error": "No se pudo identificar el edificio. Asegúrate de que una de las pestañas se llame 'Carmen 339' o 'Tarapacá 1140'."
            }), 400

        # 4. Actualizar Matriz
        actualizados_matriz = 0
        for numero_depto, estado in estados_departamentos.items():
            piso, columna = parsear_numero_departamento(numero_depto)
            
            if not piso or not columna:
                continue
            
            # Validar rangos según config
            if piso < config["MIN_PISO"] or piso > config["MAX_PISO"]:
                continue
            if columna < 1 or columna > config["MAX_COL"]:
                continue
            
            # Calcular fila: FILA_PISO_9 sube/baja según el piso
            fila_matriz = config["FILA_PISO_9"] + (9 - piso)
            # Calcular columna: COL_COLUMNA_1 es el inicio (C=3)
            col_num = 2 + columna # ya que C=3 y columna 1 debe ser C
            
            celda = ws.cell(row=fila_matriz, column=col_num)
            valor_actual = celda.value

            # Protección celdas especiales (SDV, PILOTO)
            if valor_actual and str(valor_actual).strip().upper() in ["SDV", "PILOTO"]:
                continue

            # Actualizar si no es Disponible y cambió el valor
            if estado.strip().upper() != "DISPONIBLE":
                if str(valor_actual).strip() != estado.strip():
                    celda.value = estado
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                    actualizados_matriz += 1

        # 5. Actualizar Comentarios
        actualizados_comentarios = 0
        col_nro_letra = config["COL_NRO_UNIDAD_COM"]
        col_est_letra = config["COL_ESTATUS_COM"]
        fila_com = 4
        
        while fila_com < 1000: # Límite seguridad
            celda_u = ws[f"{col_nro_letra}{fila_com}"]
            val_u = celda_u.value
            
            if val_u is None or val_u == "":
                break
            
            try:
                num_u = int(float(val_u))
                if num_u in estados_departamentos:
                    estado_nuevo = estados_departamentos[num_u]
                    if estado_nuevo.strip().upper() != "DISPONIBLE":
                        celda_e = ws[f"{col_est_letra}{fila_com}"]
                        if str(celda_e.value).strip() != estado_nuevo.strip():
                            celda_e.value = estado_nuevo
                            celda_e.alignment = Alignment(horizontal='center', vertical='center')
                            actualizados_comentarios += 1
            except:
                pass
            fila_com += 1

        # 6. Guardar en memoria y retornar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Estatus_Actualizado.xlsx"
        )

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno: {str(e)}"}), 500

# Punto de entrada para ejecución local (opcional)
if __name__ == '__main__':
    app.run(debug=True, port=5000)
